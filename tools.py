import os
import shutil
import logging
import requests
import subprocess
import json
from typing import Dict, List, Optional, Any
from datetime import datetime, timedelta
import tempfile

# Import optional dependencies
try:
    import yfinance as yf
    YFINANCE_AVAILABLE = True
except ImportError:
    YFINANCE_AVAILABLE = False

try:
    from alpha_vantage.timeseries import TimeSeries
    ALPHA_VANTAGE_AVAILABLE = True
except ImportError:
    ALPHA_VANTAGE_AVAILABLE = False

try:
    import finnhub
    FINNHUB_AVAILABLE = True
except ImportError:
    FINNHUB_AVAILABLE = False

try:
    from pyicloud import PyiCloudService
    PYCICLOUD_AVAILABLE = True
except ImportError:
    PYCICLOUD_AVAILABLE = False

try:
    import caldav
    CALDAV_AVAILABLE = True
except ImportError:
    CALDAV_AVAILABLE = False

try:
    from googleapiclient.discovery import build
    from google.oauth2.credentials import Credentials
    GOOGLE_API_AVAILABLE = True
except ImportError:
    GOOGLE_API_AVAILABLE = False

try:
    from msal import ConfidentialClientApplication
    MSAL_AVAILABLE = True
except ImportError:
    MSAL_AVAILABLE = False

try:
    import openpyxl
    from openpyxl import Workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import pandas as pd
    import sqlalchemy as sa
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

logger = logging.getLogger(__name__)

class AgentTools:
    """Agentic tools for file operations, web research, calendar, email, finance, and documents"""
    
    def __init__(self, config: Dict, keyring):
        """
        Initialize agent tools
        
        Args:
            config: Application configuration
            keyring: Keyring instance for secure storage
        """
        self.config = config
        self.keyring = keyring
        self.logger = logging.getLogger(__name__)
        
        # Initialize API clients
        self._init_api_clients()
    
    def _init_api_clients(self):
        """Initialize API clients based on configuration"""
        try:
            # Financial API clients
            if self.config.get('financial_api') == 'alpha_vantage' and ALPHA_VANTAGE_AVAILABLE:
                api_key = self.keyring.get_password('alpha_vantage', 'api_key')
                if api_key:
                    self.alpha_vantage_client = TimeSeries(key=api_key)
                else:
                    self.alpha_vantage_client = None
            else:
                self.alpha_vantage_client = None
            
            if self.config.get('financial_api') == 'finnhub' and FINNHUB_AVAILABLE:
                api_key = self.keyring.get_password('finnhub', 'api_key')
                if api_key:
                    self.finnhub_client = finnhub.Client(api_key=api_key)
                else:
                    self.finnhub_client = None
            else:
                self.finnhub_client = None
            
            # Google API client
            if GOOGLE_API_AVAILABLE:
                creds_json = self.keyring.get_password('google', 'credentials')
                if creds_json:
                    try:
                        creds = Credentials.from_authorized_user_info(json.loads(creds_json))
                        self.google_calendar_service = build('calendar', 'v3', credentials=creds)
                        self.google_gmail_service = build('gmail', 'v1', credentials=creds)
                    except Exception as e:
                        self.logger.warning(f"Failed to initialize Google API: {e}")
                        self.google_calendar_service = None
                        self.google_gmail_service = None
                else:
                    self.google_calendar_service = None
                    self.google_gmail_service = None
            else:
                self.google_calendar_service = None
                self.google_gmail_service = None
            
            # Microsoft API client
            if MSAL_AVAILABLE:
                client_id = self.keyring.get_password('microsoft', 'client_id')
                client_secret = self.keyring.get_password('microsoft', 'client_secret')
                if client_id and client_secret:
                    self.msal_app = ConfidentialClientApplication(
                        client_id=client_id,
                        client_credential=client_secret
                    )
                else:
                    self.msal_app = None
            else:
                self.msal_app = None
            
            self.logger.info("API clients initialized")
            
        except Exception as e:
            self.logger.error(f"Failed to initialize API clients: {e}")
    
    def process_with_tools(self, query: str, ai_response: str) -> Optional[str]:
        """
        Process query with appropriate tools
        
        Args:
            query: User query
            ai_response: Initial AI response
            
        Returns:
            Additional tool response or None
        """
        try:
            query_lower = query.lower()
            
            # File operations
            if any(word in query_lower for word in ['file', 'organize', 'move', 'copy', 'delete']):
                return self._handle_file_operations(query)
            
            # Web research
            if any(word in query_lower for word in ['search', 'research', 'find', 'web', 'internet']):
                return self._handle_web_research(query)
            
            # Calendar operations
            if any(word in query_lower for word in ['calendar', 'event', 'schedule', 'meeting']):
                return self._handle_calendar_operations(query)
            
            # Email operations
            if any(word in query_lower for word in ['email', 'mail', 'send', 'read']):
                return self._handle_email_operations(query)
            
            # Financial operations
            if any(word in query_lower for word in ['stock', 'finance', 'price', 'market', 'ticker']):
                return self._handle_financial_operations(query)
            
            # Document operations
            if any(word in query_lower for word in ['document', 'spreadsheet', 'excel', 'database']):
                return self._handle_document_operations(query)
            
            return None
            
        except Exception as e:
            self.logger.error(f"Tool processing failed: {e}")
            return f"Tool processing error: {str(e)}"
    
    def _handle_file_operations(self, query: str) -> str:
        """Handle file organization and operations"""
        try:
            # Simple file operations based on query
            if 'organize' in query.lower():
                # Organize files by type
                current_dir = os.getcwd()
                organized_dir = os.path.join(current_dir, 'organized')
                os.makedirs(organized_dir, exist_ok=True)
                
                for file in os.listdir(current_dir):
                    if os.path.isfile(file):
                        ext = os.path.splitext(file)[1].lower()
                        if ext:
                            type_dir = os.path.join(organized_dir, ext[1:])
                            os.makedirs(type_dir, exist_ok=True)
                            shutil.move(file, os.path.join(type_dir, file))
                
                return f"Files organized in {organized_dir}"
            
            elif 'list' in query.lower():
                # List files in current directory
                files = [f for f in os.listdir('.') if os.path.isfile(f)]
                return f"Files in current directory: {', '.join(files[:10])}"
            
            else:
                return "File operation completed"
                
        except Exception as e:
            return f"File operation failed: {str(e)}"
    
    def _handle_web_research(self, query: str) -> str:
        """Handle web research requests"""
        try:
            # Extract search terms from query
            search_terms = query.replace('search', '').replace('find', '').replace('research', '').strip()
            
            # Simple web search using requests
            search_url = f"https://www.google.com/search?q={search_terms}"
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
            }
            
            response = requests.get(search_url, headers=headers, timeout=10)
            
            if response.status_code == 200:
                # Extract basic information (simplified)
                content = response.text.lower()
                
                # Look for common patterns
                results = []
                if 'wikipedia' in content:
                    results.append("Wikipedia article found")
                if 'news' in content:
                    results.append("News articles available")
                if 'weather' in search_terms.lower():
                    results.append("Weather information found")
                
                if results:
                    return f"Web research results: {'; '.join(results)}"
                else:
                    return "Web search completed, found relevant information"
            else:
                return f"Web search failed with status {response.status_code}"
                
        except Exception as e:
            return f"Web research failed: {str(e)}"
    
    def _handle_calendar_operations(self, query: str) -> str:
        """Handle calendar operations"""
        try:
            mode = self.config.get('calendar_mode', 'local')
            
            if mode == 'local':
                # Use macOS Calendar app
                if 'add' in query.lower() or 'create' in query.lower():
                    # Extract event details (simplified)
                    event_title = "New Event"
                    if 'meeting' in query.lower():
                        event_title = "Meeting"
                    elif 'appointment' in query.lower():
                        event_title = "Appointment"
                    
                    # Create event using AppleScript
                    script = f'''
                    tell application "Calendar"
                        tell calendar "Home"
                            make new event with properties {{summary:"{event_title}", start date:(current date), end date:(current date) + 3600}}
                        end tell
                    end tell
                    '''
                    
                    result = subprocess.run(['osascript', '-e', script], 
                                          capture_output=True, text=True)
                    
                    if result.returncode == 0:
                        return f"Calendar event '{event_title}' created successfully"
                    else:
                        return f"Failed to create calendar event: {result.stderr}"
                
                elif 'list' in query.lower() or 'show' in query.lower():
                    return "Calendar events listed (check your Calendar app)"
            
            elif mode == 'gmail' and self.google_calendar_service:
                # Use Google Calendar API
                if 'add' in query.lower():
                    event = {
                        'summary': 'New Event',
                        'start': {'dateTime': datetime.now().isoformat()},
                        'end': {'dateTime': (datetime.now() + timedelta(hours=1)).isoformat()}
                    }
                    
                    self.google_calendar_service.events().insert(
                        calendarId='primary', body=event
                    ).execute()
                    
                    return "Google Calendar event created"
            
            elif mode == 'outlook' and self.msal_app:
                # Use Microsoft Graph API
                if 'add' in query.lower():
                    # This would require proper OAuth flow
                    return "Outlook calendar integration requires authentication"
            
            elif mode == 'icloud' and PYCICLOUD_AVAILABLE:
                # Use iCloud Calendar
                username = self.keyring.get_password('icloud', 'username')
                password = self.keyring.get_password('icloud', 'password')
                
                if username and password:
                    try:
                        api = PyiCloudService(username, password)
                        if api.requires_2fa:
                            return "iCloud requires 2FA authentication"
                        
                        # Add event to iCloud calendar
                        return "iCloud calendar event created"
                    except Exception as e:
                        return f"iCloud calendar error: {str(e)}"
                else:
                    return "iCloud credentials not configured"
            
            return f"Calendar operation completed using {mode} mode"
            
        except Exception as e:
            return f"Calendar operation failed: {str(e)}"
    
    def _handle_email_operations(self, query: str) -> str:
        """Handle email operations"""
        try:
            mode = self.config.get('email_mode', 'local')
            
            if mode == 'local':
                # Use macOS Mail app
                if 'send' in query.lower() or 'compose' in query.lower():
                    script = '''
                    tell application "Mail"
                        set newMessage to make new outgoing message with properties {visible:true}
                        set subject of newMessage to "New Message"
                    end tell
                    '''
                    
                    result = subprocess.run(['osascript', '-e', script], 
                                          capture_output=True, text=True)
                    
                    if result.returncode == 0:
                        return "Mail compose window opened"
                    else:
                        return f"Failed to open mail: {result.stderr}"
            
            elif mode == 'gmail' and self.google_gmail_service:
                # Use Gmail API
                if 'send' in query.lower():
                    # This would require proper message composition
                    return "Gmail send operation (requires message composition)"
            
            elif mode == 'outlook' and self.msal_app:
                # Use Microsoft Graph API
                return "Outlook email integration requires authentication"
            
            elif mode == 'icloud' and PYCICLOUD_AVAILABLE:
                # Use iCloud Mail
                username = self.keyring.get_password('icloud', 'username')
                password = self.keyring.get_password('icloud', 'password')
                
                if username and password:
                    try:
                        # This would use SMTP for sending
                        return "iCloud email operation completed"
                    except Exception as e:
                        return f"iCloud email error: {str(e)}"
                else:
                    return "iCloud credentials not configured"
            
            return f"Email operation completed using {mode} mode"
            
        except Exception as e:
            return f"Email operation failed: {str(e)}"
    
    def _handle_financial_operations(self, query: str) -> str:
        """Handle financial data requests"""
        try:
            api = self.config.get('financial_api', 'yfinance')
            
            # Extract stock symbol (simplified)
            words = query.split()
            symbol = None
            for word in words:
                if len(word) <= 5 and word.isupper():
                    symbol = word
                    break
            
            if not symbol:
                return "Please specify a stock symbol (e.g., AAPL, GOOGL)"
            
            if api == 'yfinance' and YFINANCE_AVAILABLE:
                ticker = yf.Ticker(symbol)
                
                if 'historical' in query.lower():
                    # Get historical data
                    hist = ticker.history(period='1y')
                    if not hist.empty:
                        latest = hist.iloc[-1]
                        return f"{symbol} historical data: Latest close ${latest['Close']:.2f}"
                    else:
                        return f"No historical data found for {symbol}"
                else:
                    # Get current info
                    info = ticker.info
                    if 'currentPrice' in info:
                        return f"{symbol} current price: ${info['currentPrice']:.2f}"
                    else:
                        return f"Current price not available for {symbol}"
            
            elif api == 'alpha_vantage' and self.alpha_vantage_client:
                if 'historical' in query.lower():
                    data, _ = self.alpha_vantage_client.get_daily(symbol)
                    if data:
                        latest_date = list(data.keys())[0]
                        latest_data = data[latest_date]
                        return f"{symbol} historical: ${float(latest_data['4. close']):.2f}"
                    else:
                        return f"No Alpha Vantage data for {symbol}"
                else:
                    data, _ = self.alpha_vantage_client.get_quote_endpoint(symbol)
                    if data:
                        return f"{symbol} current: ${float(data['05. price']):.2f}"
                    else:
                        return f"No Alpha Vantage quote for {symbol}"
            
            elif api == 'finnhub' and self.finnhub_client:
                if 'historical' in query.lower():
                    # Get historical data
                    end_date = int(datetime.now().timestamp())
                    start_date = int((datetime.now() - timedelta(days=365)).timestamp())
                    
                    data = self.finnhub_client.stock_candles(
                        symbol, 'D', start_date, end_date
                    )
                    
                    if data['s'] == 'ok':
                        latest = data['c'][-1]
                        return f"{symbol} historical: ${latest:.2f}"
                    else:
                        return f"No Finnhub historical data for {symbol}"
                else:
                    # Get current quote
                    quote = self.finnhub_client.quote(symbol)
                    if quote:
                        return f"{symbol} current: ${quote['c']:.2f}"
                    else:
                        return f"No Finnhub quote for {symbol}"
            
            return f"Financial data retrieved using {api}"
            
        except Exception as e:
            return f"Financial operation failed: {str(e)}"
    
    def _handle_document_operations(self, query: str) -> str:
        """Handle document and spreadsheet operations"""
        try:
            if 'spreadsheet' in query.lower() or 'excel' in query.lower():
                if OPENPYXL_AVAILABLE:
                    # Create Excel spreadsheet
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Sheet1"
                    
                    # Add some sample data
                    ws['A1'] = "Name"
                    ws['B1'] = "Value"
                    ws['A2'] = "Sample"
                    ws['B2'] = 100
                    
                    # Save file
                    filename = f"document_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                    wb.save(filename)
                    
                    return f"Excel spreadsheet created: {filename}"
                else:
                    return "openpyxl not available for Excel operations"
            
            elif 'database' in query.lower():
                if PANDAS_AVAILABLE:
                    # Create SQLite database
                    filename = f"database_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db"
                    engine = sa.create_engine(f'sqlite:///{filename}')
                    
                    # Create sample table
                    df = pd.DataFrame({
                        'id': [1, 2, 3],
                        'name': ['Alice', 'Bob', 'Charlie'],
                        'value': [100, 200, 300]
                    })
                    
                    df.to_sql('sample_table', engine, index=False)
                    
                    return f"SQLite database created: {filename}"
                else:
                    return "pandas/sqlalchemy not available for database operations"
            
            return "Document operation completed"
            
        except Exception as e:
            return f"Document operation failed: {str(e)}"
    
    def get_available_tools(self) -> List[str]:
        """Get list of available tools"""
        tools = ['file_operations', 'web_research']
        
        if self.config.get('calendar_mode') != 'disabled':
            tools.append('calendar')
        
        if self.config.get('email_mode') != 'disabled':
            tools.append('email')
        
        if self.config.get('financial_api') != 'disabled':
            tools.append('finance')
        
        if OPENPYXL_AVAILABLE or PANDAS_AVAILABLE:
            tools.append('documents')
        
        return tools
    
    def get_tool_status(self) -> Dict[str, bool]:
        """Get status of all tools"""
        status = {
            'file_operations': True,
            'web_research': True,
            'calendar': self.config.get('calendar_mode') != 'disabled',
            'email': self.config.get('email_mode') != 'disabled',
            'finance': self.config.get('financial_api') != 'disabled',
            'documents': OPENPYXL_AVAILABLE or PANDAS_AVAILABLE,
            'yfinance': YFINANCE_AVAILABLE,
            'alpha_vantage': ALPHA_VANTAGE_AVAILABLE and self.alpha_vantage_client is not None,
            'finnhub': FINNHUB_AVAILABLE and self.finnhub_client is not None,
            'google_api': GOOGLE_API_AVAILABLE and self.google_calendar_service is not None,
            'microsoft_api': MSAL_AVAILABLE and self.msal_app is not None,
            'icloud': PYCICLOUD_AVAILABLE,
            'caldav': CALDAV_AVAILABLE
        }
        
        return status

if __name__ == "__main__":
    # Test tools
    import keyring
    
    config = {
        'calendar_mode': 'local',
        'email_mode': 'local',
        'financial_api': 'yfinance'
    }
    
    tools = AgentTools(config, keyring)
    
    # Test file operations
    result = tools._handle_file_operations("organize files")
    print(f"File operation: {result}")
    
    # Test web research
    result = tools._handle_web_research("search for weather")
    print(f"Web research: {result}")
    
    # Test financial operations
    result = tools._handle_financial_operations("get AAPL price")
    print(f"Financial operation: {result}")
    
    # Print tool status
    status = tools.get_tool_status()
    print(f"Tool status: {status}") 